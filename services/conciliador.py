from __future__ import annotations

import csv
import io
import site
import sys
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

user_site = site.getusersitepackages()
if isinstance(user_site, str) and user_site and user_site not in sys.path:
    sys.path.append(user_site)

from app.models import (
    ConfiancaVinculo,
    ImportacaoExtrato,
    LancamentoComponente,
    RegraConciliador,
    StatusImportacao,
    StatusVinculoTarifa,
    TipoArquivo,
    TipoComponenteLancamento,
    TipoComparacao,
    TipoLancamento,
    TipoMovimento,
    TransacaoImportada,
)


MONTH_LABELS = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
DEFAULT_NORMALIZATION_OPTIONS = {
    "remover_numeros": False,
    "remover_especiais": True,
    "remover_acentos": True,
    "maiusculo": True,
    "colapsar_espacos": True,
}


@dataclass(slots=True)
class ParsedTransaction:
    linha_origem: int | None
    data_movimento: date
    data_ocorrencia: date | None
    descricao_original: str
    descricao_normalizada: str
    valor: Decimal
    tipo_movimento: str
    tipo_lancamento: str
    dados_brutos: dict[str, Any]


def detect_tipo_arquivo(filename: str) -> str:
    extension = Path(filename or "").suffix.lower()
    mapping = {
        ".csv": TipoArquivo.CSV,
        ".xls": TipoArquivo.XLS,
        ".xlsx": TipoArquivo.XLSX,
        ".pdf": TipoArquivo.PDF,
    }
    if extension not in mapping:
        raise ValueError("Arquivo inválido. Use CSV, XLS, XLSX ou PDF.")
    return mapping[extension]


def normalize_text(value: Any, *, options: dict[str, Any] | None = None) -> str:
    merged = {**DEFAULT_NORMALIZATION_OPTIONS, **(options or {})}
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    if merged.get("remover_acentos", True):
        text = "".join(char for char in text if not unicodedata.combining(char))

    text = re.sub(r"[-–—_/\\|]+", " ", text)

    if merged.get("remover_numeros", True):
        text = re.sub(r"\d+", " ", text)

    if merged.get("remover_especiais", True):
        text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)

    if merged.get("colapsar_espacos", True):
        text = re.sub(r"\s+", " ", text)

    text = text.strip()
    if merged.get("maiusculo", True):
        text = text.upper()
    return text


def _read_uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.open("rb")
    try:
        return uploaded_file.read()
    finally:
        uploaded_file.close()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _normalize_key(value: Any) -> str:
    return normalize_text(value, options={"remover_numeros": False, "remover_especiais": True, "maiusculo": True})


def _get_row_value(row: dict[str, Any], wanted_key: str | None) -> Any:
    if not wanted_key:
        return ""

    normalized_wanted = _normalize_key(wanted_key)
    for key, value in row.items():
        if _normalize_key(key) == normalized_wanted:
            return value

    return row.get(wanted_key, "")


def _parse_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return Decimal("0")

    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("R$", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.-]", "", text)

    if not text or text in {"-", "."}:
        return Decimal("0")

    try:
        amount = Decimal(text)
    except InvalidOperation:
        return Decimal("0")

    if negative or amount < 0:
        return abs(amount)

    return amount


def _value_is_negative(value: Any) -> bool:
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)) < 0

    text = str(value or "").strip()
    return text.startswith("-") or (text.startswith("(") and text.endswith(")"))


def _movement_from_indicator(value: Any) -> str | None:
    indicator = normalize_text(value, options={"remover_numeros": False, "remover_especiais": True})
    if indicator in {"D", "DEBITO"}:
        return TipoMovimento.DEBITO
    if indicator in {"C", "CREDITO"}:
        return TipoMovimento.CREDITO
    return None


def _reference_date_from_importacao(importacao: ImportacaoExtrato) -> date | None:
    if not importacao.referencia:
        return None

    try:
        return datetime.strptime(f"{importacao.referencia}-01", "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_date(value: Any, *, date_format: str | None = None, reference_date: date | None = None) -> date:
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    if not text:
        raise ValueError("Data não encontrada na linha do extrato.")

    formats = []
    if date_format:
        formats.append(date_format)
    formats.extend(["%d/%m/%Y", "%d/%m/%y", "%d/%m/%Y %H:%M", "%d/%m/%y %H:%M", "%d/%m"])

    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%d/%m" and reference_date:
                parsed = parsed.replace(year=reference_date.year)
            return parsed.date()
        except ValueError:
            continue

    raise ValueError(f"Não foi possível interpretar a data '{text}'.")


def _extract_occurrence_date(description: str, *, reference_date: date | None = None) -> date | None:
    text = str(description or "")
    patterns = [
        r"OCORR(?:E|Ê)NCIA\s*(\d{2}/\d{2}/\d{2,4})",
        r"OCORR\.?\s*(\d{2}/\d{2}/\d{2,4})",
        r"TAR\.?\s+AGRUPADAS\s*[-:]?\s*OCORR(?:E|Ê)NCIA\s*(\d{2}/\d{2}/\d{2,4})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        try:
            return _parse_date(match.group(1), reference_date=reference_date)
        except ValueError:
            continue
    return None


def _classify_transaction(parsed: ParsedTransaction) -> tuple[str, date | None]:
    description = parsed.descricao_normalizada or normalize_text(parsed.descricao_original)
    occurrence_date = _extract_occurrence_date(parsed.descricao_original, reference_date=parsed.data_movimento) or parsed.data_movimento

    grouped_patterns = [
        "DEBITO SERVICO COBRANCA",
        "DEBITO SERVICO COBRANCA",
        "TAR AGRUPADAS",
        "TARIFAS AGRUPADAS",
    ]
    if any(pattern in description for pattern in grouped_patterns):
        return TipoLancamento.TARIFA_AGRUPADA, occurrence_date

    tariff_patterns = [
        "TARIFA",
        "PACOTE SERVICOS",
        "PACOTE DE SERVICOS",
        "MANUTENCAO PJ",
        "MENSALIDADE PACOTE",
        "SERVICO COBRANCA",
    ]
    if any(pattern in description for pattern in tariff_patterns):
        return TipoLancamento.TARIFA, occurrence_date

    return TipoLancamento.PRINCIPAL, None


def _sync_default_component(transaction_obj: TransacaoImportada) -> None:
    if transaction_obj.tipo_lancamento != TipoLancamento.PRINCIPAL:
        transaction_obj.componentes.all().delete()
        return

    component, _created = LancamentoComponente.objects.get_or_create(
        lancamento=transaction_obj,
        tipo_componente=TipoComponenteLancamento.PRINCIPAL,
        defaults={
            "valor": transaction_obj.valor,
            "descricao": transaction_obj.descricao_original[:255],
        },
    )
    changed = False
    if component.valor != transaction_obj.valor:
        component.valor = transaction_obj.valor
        changed = True
    descricao = transaction_obj.descricao_original[:255]
    if component.descricao != descricao:
        component.descricao = descricao
        changed = True
    if changed:
        component.save(update_fields=["valor", "descricao", "atualizado_em"])


def describe_transaction_metadata(
    descricao_original: str,
    *,
    descricao_normalizada: str = "",
    data_movimento: date | None = None,
) -> dict[str, Any]:
    normalized = descricao_normalizada or normalize_text(descricao_original)
    fallback_date = data_movimento or timezone.localdate()
    parsed = ParsedTransaction(
        linha_origem=None,
        data_movimento=fallback_date,
        data_ocorrencia=None,
        descricao_original=descricao_original,
        descricao_normalizada=normalized,
        valor=Decimal("0"),
        tipo_movimento=TipoMovimento.AMBOS,
        tipo_lancamento=TipoLancamento.PRINCIPAL,
        dados_brutos={},
    )
    tipo_lancamento, data_ocorrencia = _classify_transaction(parsed)
    return {
        "tipo_lancamento": tipo_lancamento,
        "data_ocorrencia": data_ocorrencia,
    }


def _extract_csv_rows(uploaded_file) -> tuple[list[str], list[dict[str, Any]]]:
    content = _decode_text(_read_uploaded_bytes(uploaded_file))
    sample = content[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(content), dialect=dialect)
    headers = [header for header in (reader.fieldnames or []) if header]
    rows = [dict(row) for row in reader]
    return headers, rows


def _extract_xlsx_rows(uploaded_file) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError("Dependência para XLS/XLSX ausente. Instale openpyxl.") from exc

    workbook = load_workbook(filename=io.BytesIO(_read_uploaded_bytes(uploaded_file)), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    first_row = next(iterator, ())
    headers = [str(value).strip() if value is not None else "" for value in first_row]
    if not any(headers):
        headers = [f"COLUNA_{index + 1}" for index in range(len(first_row))]

    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = {}
        for index, header in enumerate(headers):
            row[header] = values[index] if index < len(values) else ""
        rows.append(row)

    return headers, rows


def _extract_pdf_text(uploaded_file) -> tuple[list[str], list[dict[str, Any]], str]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError("Dependência para PDF ausente. Instale pypdf.") from exc

    reader = PdfReader(io.BytesIO(_read_uploaded_bytes(uploaded_file)))
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text:
            text_parts.append(page_text)

    text = "\n".join(text_parts)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    rows: list[dict[str, Any]] = []
    pattern = re.compile(
        r"(?P<data>\d{2}/\d{2}(?:/\d{2,4})?)\s+(?P<descricao>.+?)\s+(?P<valor>\(?-?[\d.]+,[\d]{2}\)?)(?:\s+(?P<tipo>[CD]))?$",
        flags=re.IGNORECASE,
    )
    for line_number, line in enumerate(lines, start=1):
        match = pattern.search(line)
        if not match:
            continue

        rows.append(
            {
                "DATA": match.group("data"),
                "DESCRICAO": match.group("descricao"),
                "VALOR": match.group("valor"),
                "TIPO": match.group("tipo") or "",
                "__line__": line_number,
            }
        )

    return lines, rows, text


def _resolve_mapped_column(mapping: dict[str, Any], key: str, headers: list[str]) -> str:
    value = mapping.get(key)
    if value:
        return str(value)

    aliases = {
        "data": ["DATA", "DATA_MOVIMENTO", "DATA LANÇAMENTO", "DATA LANCAMENTO", "MOVIMENTO", "DATE"],
        "descricao": ["DESCRICAO", "HISTORICO", "HISTORICO LANCAMENTO", "LANÇAMENTO", "LANCAMENTO", "DESCRIPTION"],
        "valor": ["VALOR", "AMOUNT", "VLR", "DEBITO", "CREDITO"],
        "credito": ["CREDITO", "CRÉDITO", "CREDIT"],
        "debito": ["DEBITO", "DÉBITO", "DEBIT"],
    }

    normalized_headers = { _normalize_key(header): header for header in headers }
    for alias in aliases.get(key, []):
        alias_key = _normalize_key(alias)
        if alias_key in normalized_headers:
            return normalized_headers[alias_key]
    return ""


def _build_transaction_row(
    row: dict[str, Any],
    *,
    linha_origem: int | None,
    importacao: ImportacaoExtrato,
    headers: list[str],
    configuracao: dict[str, Any],
) -> ParsedTransaction:
    colunas = configuracao.get("colunas") or {}
    date_format = configuracao.get("data_format") or configuracao.get("formato_data") or ""
    normalizacao = configuracao.get("normalizacao") or {}
    referencia = _reference_date_from_importacao(importacao)

    data_column = _resolve_mapped_column(colunas, "data", headers)
    descricao_column = _resolve_mapped_column(colunas, "descricao", headers)
    valor_column = _resolve_mapped_column(colunas, "valor", headers)
    credito_column = _resolve_mapped_column(colunas, "credito", headers)
    debito_column = _resolve_mapped_column(colunas, "debito", headers)

    raw_data = _get_row_value(row, data_column)
    raw_descricao = _get_row_value(row, descricao_column)
    raw_valor = _get_row_value(row, valor_column)
    raw_credito = _get_row_value(row, credito_column)
    raw_debito = _get_row_value(row, debito_column)
    raw_tipo = _get_row_value(row, "TIPO") or _get_row_value(row, "TIPO_MOVIMENTO")

    descricao_original = str(raw_descricao or raw_valor or "").strip()
    if not descricao_original:
        descricao_original = str(row.get("DESCRICAO") or row.get("DESCRIÇÃO") or row.get("HISTORICO") or "").strip()

    if raw_credito not in {None, ""}:
        tipo_movimento = TipoMovimento.CREDITO
        valor = _parse_decimal(raw_credito)
    elif raw_debito not in {None, ""}:
        tipo_movimento = TipoMovimento.DEBITO
        valor = _parse_decimal(raw_debito)
    else:
        valor = _parse_decimal(raw_valor)
        tipo_movimento = _movement_from_indicator(raw_tipo)
        if not tipo_movimento:
            tipo_movimento = TipoMovimento.DEBITO if _value_is_negative(raw_valor) else TipoMovimento.CREDITO

    data_movimento = _parse_date(raw_data, date_format=date_format or None, reference_date=referencia)
    descricao_normalizada = normalize_text(descricao_original, options=normalizacao or None)

    provisional = ParsedTransaction(
        linha_origem=linha_origem,
        data_movimento=data_movimento,
        data_ocorrencia=None,
        descricao_original=descricao_original,
        descricao_normalizada=descricao_normalizada,
        valor=abs(valor),
        tipo_movimento=tipo_movimento,
        tipo_lancamento=TipoLancamento.PRINCIPAL,
        dados_brutos={
            **row,
            "__linha_origem__": linha_origem,
        },
    )
    tipo_lancamento, data_ocorrencia = _classify_transaction(provisional)

    return ParsedTransaction(
        linha_origem=linha_origem,
        data_movimento=data_movimento,
        data_ocorrencia=data_ocorrencia,
        descricao_original=descricao_original,
        descricao_normalizada=descricao_normalizada,
        valor=abs(valor),
        tipo_movimento=tipo_movimento,
        tipo_lancamento=tipo_lancamento,
        dados_brutos={
            **row,
            "__linha_origem__": linha_origem,
        },
    )


def inspect_importacao_file(importacao: ImportacaoExtrato) -> dict[str, Any]:
    if importacao.tipo_arquivo == TipoArquivo.CSV:
        headers, rows = _extract_csv_rows(importacao.arquivo)
        return {
            "cabecalhos": headers,
            "amostra": rows[:5],
            "tipo": importacao.tipo_arquivo,
        }

    if importacao.tipo_arquivo in {TipoArquivo.XLS, TipoArquivo.XLSX}:
        headers, rows = _extract_xlsx_rows(importacao.arquivo)
        return {
            "cabecalhos": headers,
            "amostra": rows[:5],
            "tipo": importacao.tipo_arquivo,
        }

    linhas, rows, text = _extract_pdf_text(importacao.arquivo)
    return {
        "linhas_extraidas": len(linhas),
        "amostra": rows[:5],
        "texto_preview": text[:4000],
        "tipo": importacao.tipo_arquivo,
    }


def _parse_importacao_rows(importacao: ImportacaoExtrato, configuracao: dict[str, Any]) -> tuple[list[ParsedTransaction], dict[str, Any]]:
    configuracao = configuracao or {}
    if importacao.tipo_arquivo == TipoArquivo.CSV:
        headers, rows = _extract_csv_rows(importacao.arquivo)
        parsed = [
            _build_transaction_row(row, linha_origem=index, importacao=importacao, headers=headers, configuracao=configuracao)
            for index, row in enumerate(rows, start=1)
            if any(str(value).strip() for value in row.values())
        ]
        return parsed, {"cabecalhos": headers, "amostra": rows[:5], "tipo": importacao.tipo_arquivo}

    if importacao.tipo_arquivo in {TipoArquivo.XLS, TipoArquivo.XLSX}:
        headers, rows = _extract_xlsx_rows(importacao.arquivo)
        parsed = [
            _build_transaction_row(row, linha_origem=index, importacao=importacao, headers=headers, configuracao=configuracao)
            for index, row in enumerate(rows, start=1)
            if any(str(value).strip() for value in row.values())
        ]
        return parsed, {"cabecalhos": headers, "amostra": rows[:5], "tipo": importacao.tipo_arquivo}

    linhas, rows, text = _extract_pdf_text(importacao.arquivo)
    parsed = []
    for index, row in enumerate(rows, start=1):
        parsed.append(_build_transaction_row(row, linha_origem=index, importacao=importacao, headers=list(row.keys()), configuracao=configuracao))
    return parsed, {
        "linhas_extraidas": len(linhas),
        "amostra": rows[:5],
        "texto_preview": text[:4000],
        "tipo": importacao.tipo_arquivo,
    }


def process_importacao(importacao: ImportacaoExtrato, configuracao: dict[str, Any] | None = None) -> dict[str, Any]:
    with transaction.atomic():
        parsed_rows, metadata = _parse_importacao_rows(importacao, configuracao or {})
        if not parsed_rows:
            importacao.status = StatusImportacao.ERRO
            importacao.mensagem_erro = "Nenhuma transação válida foi encontrada no arquivo."
            importacao.metadados = metadata
            importacao.save(update_fields=["status", "mensagem_erro", "metadados", "atualizado_em"])
            raise ValueError(importacao.mensagem_erro)

        existing = {transaction.linha_origem: transaction for transaction in importacao.transacoes.all()}
        seen_lines: set[int] = set()

        for parsed in parsed_rows:
            line_number = parsed.linha_origem
            if line_number is not None:
                seen_lines.add(line_number)

            transaction_obj = existing.get(line_number)
            defaults = {
                "data_movimento": parsed.data_movimento,
                "data_ocorrencia": parsed.data_ocorrencia,
                "descricao_original": parsed.descricao_original,
                "descricao_normalizada": parsed.descricao_normalizada,
                "valor": parsed.valor,
                "tipo_movimento": parsed.tipo_movimento,
                "tipo_lancamento": parsed.tipo_lancamento,
                "lancamento_relacionado": None,
                "status_vinculo_tarifa": StatusVinculoTarifa.NAO_APLICA,
                "confianca_vinculo": ConfiancaVinculo.BAIXA,
                "dados_brutos": parsed.dados_brutos,
            }

            if transaction_obj:
                if transaction_obj.revisado_manual:
                    defaults.update(
                        {
                            "descricao_normalizada": transaction_obj.descricao_normalizada,
                            "categoria": transaction_obj.categoria,
                            "subcategoria": transaction_obj.subcategoria,
                            "conta_debito": transaction_obj.conta_debito,
                            "conta_credito": transaction_obj.conta_credito,
                            "codigo_historico": transaction_obj.codigo_historico,
                            "tipo_lancamento": transaction_obj.tipo_lancamento,
                            "data_ocorrencia": transaction_obj.data_ocorrencia,
                            "lancamento_relacionado": transaction_obj.lancamento_relacionado,
                            "status_vinculo_tarifa": transaction_obj.status_vinculo_tarifa,
                            "confianca_vinculo": transaction_obj.confianca_vinculo,
                            "regra_aplicada": transaction_obj.regra_aplicada,
                            "revisado_manual": True,
                        }
                    )
                else:
                    defaults.update(
                        {
                            "regra_aplicada": None,
                            "categoria": "",
                            "subcategoria": "",
                            "conta_debito": "",
                            "conta_credito": "",
                            "codigo_historico": "",
                            "lancamento_relacionado": None,
                            "status_vinculo_tarifa": StatusVinculoTarifa.NAO_APLICA,
                            "confianca_vinculo": ConfiancaVinculo.BAIXA,
                            "revisado_manual": False,
                        }
                    )

            transaction_obj, _created = TransacaoImportada.objects.update_or_create(
                importacao=importacao,
                linha_origem=line_number,
                defaults=defaults,
            )
            if not transaction_obj.revisado_manual:
                _sync_default_component(transaction_obj)

        if seen_lines:
            importacao.transacoes.exclude(linha_origem__in=seen_lines).filter(revisado_manual=False).delete()

        importacao.status = StatusImportacao.PROCESSADA
        importacao.mensagem_erro = ""
        importacao.metadados = metadata
        importacao.save(update_fields=["status", "mensagem_erro", "metadados", "atualizado_em"])

        return {
            "transacoes_processadas": len(parsed_rows),
            "cabecalhos": metadata.get("cabecalhos", []),
            "amostra": metadata.get("amostra", []),
            "tipo_arquivo": importacao.tipo_arquivo,
        }


def _rule_matches_description(rule: RegraConciliador, description: str) -> bool:
    reference = normalize_text(rule.texto_referencia)
    if not reference:
        return False

    target = normalize_text(description)
    if rule.tipo_comparacao == TipoComparacao.IGUAL:
        return target == reference

    if rule.tipo_comparacao == TipoComparacao.COMECA_COM:
        return target.startswith(reference)

    return reference in target


def _rule_matches_transaction(rule: RegraConciliador, transaction: TransacaoImportada) -> bool:
    if not rule.ativo or not rule.aplicar_automatico:
        return False

    if rule.empresa_id and rule.empresa_id != transaction.importacao.empresa_id:
        return False

    if rule.escritorio_id != transaction.importacao.escritorio_id:
        return False

    if rule.tipo_movimento != TipoMovimento.AMBOS and rule.tipo_movimento != transaction.tipo_movimento:
        return False

    description = transaction.descricao_normalizada or transaction.descricao_original
    return _rule_matches_description(rule, description)


def apply_rules_to_importacao(importacao: ImportacaoExtrato) -> dict[str, Any]:
    with transaction.atomic():
        rules_queryset = RegraConciliador.objects.filter(escritorio=importacao.escritorio, ativo=True).filter(
            Q(empresa__isnull=True) | Q(empresa=importacao.empresa)
        )
        rules = list(rules_queryset.order_by("prioridade", "nome", "criado_em"))

        applied = 0
        pending = 0
        manual = 0

        transacoes_para_atualizar = []

        for transaction_obj in importacao.transacoes.select_related("regra_aplicada").order_by("data_movimento", "id"):
            if transaction_obj.revisado_manual:
                manual += 1
                continue

            matched_rule = next((rule for rule in rules if _rule_matches_transaction(rule, transaction_obj)), None)

            if matched_rule:
                transaction_obj.regra_aplicada = matched_rule
                transaction_obj.categoria = matched_rule.categoria
                transaction_obj.subcategoria = matched_rule.subcategoria
                transaction_obj.conta_debito = matched_rule.conta_debito
                transaction_obj.conta_credito = matched_rule.conta_credito
                transaction_obj.codigo_historico = matched_rule.codigo_historico
                applied += 1
            else:
                transaction_obj.regra_aplicada = None
                transaction_obj.categoria = ""
                transaction_obj.subcategoria = ""
                transaction_obj.conta_debito = ""
                transaction_obj.conta_credito = ""
                transaction_obj.codigo_historico = ""
                pending += 1

            transacoes_para_atualizar.append(transaction_obj)

        if transacoes_para_atualizar:
            TransacaoImportada.objects.bulk_update(
                transacoes_para_atualizar,
                [
                    "regra_aplicada",
                    "categoria",
                    "subcategoria",
                    "conta_debito",
                    "conta_credito",
                    "codigo_historico",
                    "atualizado_em",
                ],
            )

        return {
            "regras_aplicadas": applied,
            "pendentes": pending,
            "manuais": manual,
            "regras_disponiveis": len(rules),
        }


def processar_comprovante(
    transacao: TransacaoImportada,
    comprovante_data: dict,
    *,
    usuario: str = "SISTEMA",
) -> dict[str, Any]:
    """
    Processa dados de um comprovante e cria/atualiza componentes do lançamento.

    Args:
        transacao: TransacaoImportada alvo
        comprovante_data: dict com keys:
            - tipo: "pix" | "ted" | "boleto" | "convenio"
            - valor_principal: Decimal
            - tarifa_valor: Decimal (opcional)
            - juros_valor: Decimal (opcional)
            - multa_valor: Decimal (opcional)
            - desconto_valor: Decimal (opcional)
            - documento: str (opcional)
            - data_pagamento: date (opcional)
            - beneficiario: str (opcional)
    """
    from app.models import (
        LancamentoComponente,
        OrigemAuditoriaTarifa,
        TarifaVinculoAuditoria,
    )

    tipo = comprobante_data.get("tipo", "").lower()
    valor_principal = Decimal(str(comprovante_data.get("valor_principal", 0)))
    tarifa_valor = Decimal(str(comprovante_data.get("tarifa_valor", 0)))
    juros_valor = Decimal(str(comprovante_data.get("juros_valor", 0)))
    multa_valor = Decimal(str(comprovante_data.get("multa_valor", 0)))
    desconto_valor = Decimal(str(comprovante_data.get("desconto_valor", 0)))

    if tipo in {"pix", "ted"}:
        _criar_componente_principal(transacao, valor_principal)
        if tarifa_valor > 0:
            _processar_tarifa_pix_ted(transacao, tarifa_valor, comprovante_data, usuario)
        return {
            "componentes_criados": 1,
            "tarifa_processada": tarifa_valor > 0,
            "tipo": tipo,
        }

    if tipo in {"boleto", "convenio"}:
        _criar_componentes_boleto(
            transacao,
            valor_principal,
            juros_valor,
            multa_valor,
            desconto_valor,
            comprovante_data,
        )
        return {
            "componentes_criados": 1 + sum(1 for v in [juros_valor, multa_valor, desconto_valor] if v > 0),
            "tarifa_processada": False,
            "tipo": tipo,
        }

    return {"erro": f"Tipo de comprovante não suportado: {tipo}"}


def _criar_componente_principal(transacao: TransacaoImportada, valor: Decimal) -> None:
    LancamentoComponente.objects.update_or_create(
        lancamento=transacao,
        tipo_componente=TipoComponenteLancamento.PRINCIPAL,
        defaults={
            "valor": abs(valor),
            "descricao": transacao.descricao_original[:255],
        },
    )


def _criar_componentes_boleto(
    transacao: TransacaoImportada,
    principal: Decimal,
    juros: Decimal,
    multa: Decimal,
    desconto: Decimal,
    comprovante_data: dict,
) -> None:
    LancamentoComponente.objects.update_or_create(
        lancamento=transacao,
        tipo_componente=TipoComponenteLancamento.PRINCIPAL,
        defaults={
            "valor": abs(principal),
            "descricao": f"Principal - {comprovante_data.get('beneficiario', 'Boleto')}",
        },
    )

    if juros > 0:
        LancamentoComponente.objects.update_or_create(
            lancamento=transacao,
            tipo_componente=TipoComponenteLancamento.JUROS,
            defaults={
                "valor": abs(juros),
                "descricao": "Juros de boleto",
            },
        )

    if multa > 0:
        LancamentoComponente.objects.update_or_create(
            lancamento=transacao,
            tipo_componente=TipoComponenteLancamento.MULTA,
            defaults={
                "valor": abs(multa),
                "descricao": "Multa de boleto",
            },
        )

    if desconto > 0:
        LancamentoComponente.objects.update_or_create(
            lancamento=transacao,
            tipo_componente=TipoComponenteLancamento.DESCONTO,
            defaults={
                "valor": abs(desconto),
                "descricao": "Desconto de boleto",
            },
        )

    valor_total = principal + juros + multa - desconto
    transacao.valor = abs(valor_total)
    transacao.save(update_fields=["valor", "atualizado_em"])


def _processar_tarifa_pix_ted(
    transacao: TransacaoImportada,
    tarifa_valor: Decimal,
    comprovante_data: dict,
    usuario: str,
) -> None:
    from app.models import OrigemAuditoriaTarifa, TarifaVinculoAuditoria

    status_anterior = transacao.status_vinculo_tarifa
    tarifa_existente = transacao.lancamento_relacionado

    tarifa_lancamento, created = TransacaoImportada.objects.get_or_create(
        importacao=transacao.importacao,
        tipo_lancamento=TipoLancamento.TARIFA,
        defaults={
            "data_movimento": transacao.data_movimento,
            "data_ocorrencia": transacao.data_ocorrencia or transacao.data_movimento,
            "descricao_original": f"Tarifa {transacao.descricao_original[:100]}",
            "descricao_normalizada": f"TARIFA {normalize_text(transacao.descricao_original)[:100]}",
            "valor": abs(tarifa_valor),
            "tipo_movimento": TipoMovimento.DEBITO,
            "status_vinculo_tarifa": StatusVinculoTarifa.NAO_APLICA,
            "confianca_vinculo": ConfiancaVinculo.BAIXA,
            "revisado_manual": False,
        },
    )

    if not created and tarifa_lancamento.valor != abs(tarifa_valor):
        tarifa_lancamento.valor = abs(tarifa_valor)
        tarifa_lancamento.save(update_fields=["valor", "atualizado_em"])

    transacao.lancamento_relacionado = tarifa_lancamento
    transacao.status_vinculo_tarifa = StatusVinculoTarifa.ENCONTRADA
    transacao.confianca_vinculo = ConfiancaVinculo.ALTA
    transacao.save(
        update_fields=["lancamento_relacionado", "status_vinculo_tarifa", "confianca_vinculo", "atualizado_em"]
    )

    TarifaVinculoAuditoria.objects.create(
        lancamento_principal=transacao,
        lancamento_tarifa=tarifa_lancamento,
        usuario=usuario,
        origem=OrigemAuditoriaTarifa.AUTOMATICA,
        status_anterior=status_anterior,
        status_novo=StatusVinculoTarifa.ENCONTRADA,
    )


def conciliar_tarifas_importacao(importacao: ImportacaoExtrato) -> dict[str, Any]:
    """
    Concilia tarifas para todos os lançamentos PRINCIPAL de uma importação.

    Prioridade:
    1. Tarifa PIX individual (mesma data ocorrência, mesmo valor)
    2. Tarifa agrupada (mesma data ocorrência)
    3. Não encontrada
    """
    from app.models import OrigemAuditoriaTarifa, TarifaVinculoAuditoria

    principais = list(
        importacao.transacoes.filter(
            tipo_lancamento=TipoLancamento.PRINCIPAL,
        ).select_related("lancamento_relacionado")
    )

    tarifas_individuais = list(
        importacao.transacoes.filter(
            tipo_lancamento=TipoLancamento.TARIFA,
            lancamento_relacionado__isnull=True,
        )
    )

    tarifas_agrupadas = list(
        importacao.transacoes.filter(
            tipo_lancamento=TipoLancamento.TARIFA_AGRUPADA,
            lancamento_relacionado__isnull=True,
        )
    )

    tarifas_usadas = set()

    encontradas = 0
    agrupadas = 0
    nao_encontradas = 0

    for principal in principais:
        if principal.status_vinculo_tarifa != StatusVinculoTarifa.NAO_APLICA:
            continue

        desc_norm = (principal.descricao_normalizada or "").upper()
        data_ocorrencia_principal = principal.data_ocorrencia or principal.data_movimento

        if "PIX" not in desc_norm and "TED" not in desc_norm:
            continue

        tarifa_encontrada = None
        status_novo = None
        confianca = None

        for tarifa in tarifas_individuais:
            if tarifa.id in tarifas_usadas:
                continue

            desc_tarifa = (tarifa.descricao_normalizada or "").upper()
            data_ocorrencia_tarifa = tarifa.data_ocorrencia or tarifa.data_movimento

            is_tarifa_pix = "TARIFA PIX ENVIADO" in desc_tarifa or "TARIFA PIX RECEBIDO" in desc_tarifa
            if not is_tarifa_pix:
                continue

            if data_ocorrencia_principal == data_ocorrencia_tarifa and tarifa.valor > 0:
                tarifa_encontrada = tarifa
                status_novo = StatusVinculoTarifa.ENCONTRADA
                confianca = ConfiancaVinculo.ALTA
                tarifas_usadas.add(tarifa.id)
                break

        if not tarifa_encontrada:
            for tarifa in tarifas_agrupadas:
                if tarifa.id in tarifas_usadas:
                    continue

                data_ocorrencia_tarifa = tarifa.data_ocorrencia or tarifa.data_movimento

                if data_ocorrencia_principal == data_ocorrencia_tarifa:
                    tarifa_encontrada = tarifa
                    status_novo = StatusVinculoTarifa.AGRUPADA
                    confianca = ConfiancaVinculo.MEDIA
                    tarifas_usadas.add(tarifa.id)
                    break

        if tarifa_encontrada:
            if status_novo == StatusVinculoTarifa.ENCONTRADA:
                encontradas += 1
            else:
                agrupadas += 1

            status_anterior = principal.status_vinculo_tarifa
            principal.lancamento_relacionado = tarifa_encontrada
            principal.status_vinculo_tarifa = status_novo
            principal.confianca_vinculo = confianca
            principal.save(
                update_fields=["lancamento_relacionado", "status_vinculo_tarifa", "confianca_vinculo", "atualizado_em"]
            )

            TarifaVinculoAuditoria.objects.create(
                lancamento_principal=principal,
                lancamento_tarifa=tarifa_encontrada,
                usuario="SISTEMA",
                origem=OrigemAuditoriaTarifa.AUTOMATICA,
                status_anterior=status_anterior,
                status_novo=status_novo,
            )
        else:
            nao_encontradas += 1

    return {
        "encontradas": encontradas,
        "agrupadas": agrupadas,
        "nao_encontradas": nao_encontradas,
        "total_principais": len(principais),
    }


def aplicar_vinculo_tarifa_manual(
    transacao_principal: TransacaoImportada,
    transacao_tarifa: TransacaoImportada | None,
    *,
    usuario: str = "SISTEMA",
) -> TransacaoImportada:
    """
    Vincula manualmente uma tarifa a um lançamento principal.
    """
    from app.models import OrigemAuditoriaTarifa, TarifaVinculoAuditoria

    status_anterior = transacao_principal.status_vinculo_tarifa
    tarifa_anterior_id = transacao_principal.lancamento_relacionado_id

    if transacao_tarifa:
        transacao_principal.lancamento_relacionado = transacao_tarifa
        if transacao_tarifa.tipo_lancamento == TipoLancamento.TARIFA_AGRUPADA:
            transacao_principal.status_vinculo_tarifa = StatusVinculoTarifa.AGRUPADA
            transacao_principal.confianca_vinculo = ConfiancaVinculo.MEDIA
        else:
            transacao_principal.status_vinculo_tarifa = StatusVinculoTarifa.ENCONTRADA
            transacao_principal.confianca_vinculo = ConfiancaVinculo.ALTA
    else:
        transacao_principal.lancamento_relacionado = None
        transacao_principal.status_vinculo_tarifa = StatusVinculoTarifa.NAO_ENCONTRADA
        transacao_principal.confianca_vinculo = ConfiancaVinculo.BAIXA

    transacao_principal.revisado_manual = True
    transacao_principal.save(
        update_fields=[
            "lancamento_relacionado",
            "status_vinculo_tarifa",
            "confianca_vinculo",
            "revisado_manual",
            "atualizado_em",
        ]
    )

    TarifaVinculoAuditoria.objects.create(
        lancamento_principal=transacao_principal,
        lancamento_tarifa=transacao_tarifa,
        usuario=usuario,
        origem=OrigemAuditoriaTarifa.MANUAL,
        status_anterior=status_anterior,
        status_novo=transacao_principal.status_vinculo_tarifa,
    )

    return transacao_principal
